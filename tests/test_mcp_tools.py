"""Integration tests for FastMCP tool definitions."""

import asyncio
import json
from pathlib import Path
from unittest.mock import patch

import numpy as np
import pytest
from fastmcp import Client

from server.jobs import registry
from server.main import mcp
from server.store import EMBEDDING_DIM


@pytest.fixture(autouse=True)
def clean_registry():
    """Isolate the process-wide job registry between tests, and stop it writing
    its persistence file during the run."""
    registry._jobs.clear()
    registry._order.clear()
    saved_persist = registry._persist_path
    registry._persist_path = None
    yield
    registry._persist_path = saved_persist
    registry._jobs.clear()
    registry._order.clear()


async def _wait_for_job(client, db_path, job_id, deadline=5.0):
    """Poll get_index_status until the job finishes; return its final dict.

    Bounded so a regression that hangs the job fails the test fast.
    """
    waited = 0.0
    while waited < deadline:
        status = await client.call_tool("get_index_status", {"db_path": db_path})
        data = json.loads(status.content[0].text)
        all_jobs = data["jobs"]["active"] + data["jobs"]["recent"]
        job = next((j for j in all_jobs if j["job_id"] == job_id), None)
        if job is not None and job["state"] in ("completed", "failed"):
            return job
        await asyncio.sleep(0.05)
        waited += 0.05
    raise AssertionError(f"job {job_id} did not finish within {deadline}s")


def _fake_embed(texts):
    results = []
    for t in texts:
        rng = np.random.RandomState(hash(t) % 2**32)
        vec = rng.randn(EMBEDDING_DIM).astype(np.float32)
        vec = vec / np.linalg.norm(vec)
        results.append(vec)
    return np.array(results)


@pytest.fixture
def mock_model():
    model = type("MockModel", (), {"encode": lambda self, texts, **kw: _fake_embed(texts)})()
    with patch("server.indexer.get_model", return_value=model), \
         patch("server.search.get_model", return_value=model):
        yield model


@pytest.fixture
def docs_dir(tmp_path):
    (tmp_path / "readme.md").write_text("# Revenue Report\n\nQ3 revenue grew 23% to 4.2M.")
    (tmp_path / "notes.txt").write_text("Meeting notes about the product launch.")
    return tmp_path


@pytest.mark.anyio
async def test_mcp_lists_tools():
    """MCP server exposes both tools."""
    async with Client(mcp) as client:
        tools = await client.list_tools()
        tool_names = {t.name for t in tools}
        assert "index_folder" in tool_names
        assert "semantic_search" in tool_names


@pytest.mark.anyio
async def test_mcp_index_folder_runs_as_background_job(mock_model, docs_dir, tmp_path):
    """index_folder starts a background job; polling get_index_status sees it finish."""
    db_path = str(tmp_path / "testdb")

    async with Client(mcp) as client:
        start = await client.call_tool(
            "index_folder", {"folder_path": str(docs_dir), "db_path": db_path}
        )
        assert not start.is_error
        started = json.loads(start.content[0].text)
        assert started["status"] == "started"
        job_id = started["job_id"]

        job = await _wait_for_job(client, db_path, job_id)
        assert job["state"] == "completed"
        assert job["result"]["files_indexed"] == 2


@pytest.mark.anyio
async def test_mcp_index_folder_job_fails_cleanly_on_error(docs_dir, tmp_path, monkeypatch):
    """An error raised inside the worker ends the job in 'failed', never hung."""
    def boom(*args, **kwargs):
        raise RuntimeError("indexing exploded")

    monkeypatch.setattr("server.indexer.index_folder", boom)

    db_path = str(tmp_path / "db")
    async with Client(mcp) as client:
        start = await client.call_tool(
            "index_folder", {"folder_path": str(docs_dir), "db_path": db_path}
        )
        job_id = json.loads(start.content[0].text)["job_id"]
        job = await _wait_for_job(client, db_path, job_id)

    assert job["state"] == "failed"
    assert "indexing exploded" in job["error"]
    assert registry.has_running() is False


@pytest.mark.anyio
async def test_mcp_index_folder_rejects_while_job_running(docs_dir, tmp_path):
    """A second index_folder is rejected while a job is already running."""
    running = registry.create(str(docs_dir), str(tmp_path / "db"))

    async with Client(mcp) as client:
        result = await client.call_tool(
            "index_folder",
            {"folder_path": str(docs_dir), "db_path": str(tmp_path / "db2")},
        )
        data = json.loads(result.content[0].text)
        assert data["status"] == "rejected"
        assert data["running_job_id"] == running.job_id


@pytest.mark.anyio
async def test_mcp_index_folder_missing_folder_errors(tmp_path):
    """A nonexistent folder is reported immediately, not as a background job."""
    async with Client(mcp) as client:
        result = await client.call_tool(
            "index_folder",
            {"folder_path": str(tmp_path / "nope"), "db_path": str(tmp_path / "db")},
            raise_on_error=False,
        )
        assert result.is_error
    assert registry.has_running() is False


@pytest.mark.anyio
async def test_mcp_semantic_search(mock_model, docs_dir, tmp_path):
    """semantic_search tool returns results after indexing."""
    import os
    db_path = str(tmp_path / "testdb")
    os.environ["LANCEDB_PATH"] = db_path

    from server.indexer import index_folder as _index_folder
    _index_folder(str(docs_dir), db_path=db_path)

    async with Client(mcp) as client:
        result = await client.call_tool("semantic_search", {"query": "revenue growth"})
        assert not result.is_error
        assert len(result.content) > 0
        text = result.content[0].text
        assert "results" in text


@pytest.mark.anyio
async def test_mcp_semantic_search_hybrid(mock_model, docs_dir, tmp_path):
    """semantic_search tool supports hybrid mode."""
    import os
    db_path = str(tmp_path / "testdb")
    os.environ["LANCEDB_PATH"] = db_path

    from server.indexer import index_folder as _index_folder
    _index_folder(str(docs_dir), db_path=db_path)

    async with Client(mcp) as client:
        result = await client.call_tool(
            "semantic_search", {"query": "revenue", "mode": "hybrid"}
        )
        assert not result.is_error
        text = result.content[0].text
        assert "hybrid" in text


@pytest.mark.anyio
async def test_mcp_lists_all_tools():
    """MCP server exposes all four tools."""
    async with Client(mcp) as client:
        tools = await client.list_tools()
        tool_names = {t.name for t in tools}
        assert "index_folder" in tool_names
        assert "semantic_search" in tool_names
        assert "get_index_status" in tool_names
        assert "reindex_file" in tool_names


@pytest.mark.anyio
async def test_mcp_get_index_status(mock_model, docs_dir, tmp_path):
    """get_index_status returns stats after indexing."""
    import os
    db_path = str(tmp_path / "testdb")
    os.environ["LANCEDB_PATH"] = db_path

    from server.indexer import index_folder as _index_folder
    _index_folder(str(docs_dir), db_path=db_path)

    async with Client(mcp) as client:
        result = await client.call_tool("get_index_status", {"db_path": db_path})
        assert not result.is_error
        text = result.content[0].text
        assert "total_chunks" in text
        assert "indexed_files" in text


@pytest.mark.anyio
async def test_mcp_get_index_status_empty(tmp_path):
    """get_index_status works on empty/nonexistent db."""
    db_path = str(tmp_path / "emptydb")
    async with Client(mcp) as client:
        result = await client.call_tool("get_index_status", {"db_path": db_path})
        assert not result.is_error
        text = result.content[0].text
        assert "total_chunks" in text


@pytest.mark.anyio
async def test_mcp_reindex_file(mock_model, docs_dir, tmp_path):
    """reindex_file forces re-indexing of a single file."""
    import os
    db_path = str(tmp_path / "testdb")
    os.environ["LANCEDB_PATH"] = db_path

    from server.indexer import index_folder as _index_folder
    _index_folder(str(docs_dir), db_path=db_path)

    file_path = str(docs_dir / "readme.md")
    async with Client(mcp) as client:
        result = await client.call_tool("reindex_file", {"file_path": file_path})
        assert not result.is_error
        text = result.content[0].text
        assert "reindexed" in text.lower() or "completed" in text.lower()


@pytest.mark.anyio
async def test_mcp_get_index_status_includes_jobs(tmp_path):
    """get_index_status reports a jobs section alongside the legacy keys."""
    db_path = str(tmp_path / "db")
    async with Client(mcp) as client:
        result = await client.call_tool("get_index_status", {"db_path": db_path})
        data = json.loads(result.content[0].text)
        # Legacy keys preserved for backward compatibility.
        assert "total_chunks" in data
        assert "total_files" in data
        assert "indexed_files" in data
        # New jobs section.
        assert "jobs" in data
        assert data["jobs"] == {"active": [], "recent": []}


@pytest.mark.anyio
async def test_mcp_reindex_file_rejects_while_job_running(mock_model, docs_dir, tmp_path):
    """reindex_file is rejected while a background indexing job is running."""
    db_path = str(tmp_path / "testdb")

    from server.indexer import index_folder as _index_folder
    _index_folder(str(docs_dir), db_path=db_path)

    registry.create(str(docs_dir), db_path)  # simulate a job in flight

    async with Client(mcp) as client:
        result = await client.call_tool(
            "reindex_file",
            {"file_path": str(docs_dir / "readme.md"), "db_path": db_path},
        )
        data = json.loads(result.content[0].text)
        assert data["status"] == "rejected"


@pytest.mark.anyio
async def test_mcp_reindex_file_records_stat_for_fastpath(mock_model, docs_dir, tmp_path):
    """reindex_file stores mtime/size, so a later index_folder fast-paths the file."""
    db_path = str(tmp_path / "testdb")
    from server.indexer import index_folder as _index_folder

    async with Client(mcp) as client:
        result = await client.call_tool(
            "reindex_file",
            {"file_path": str(docs_dir / "readme.md"), "db_path": db_path},
        )
        assert not result.is_error

    # readme.md was indexed via reindex_file; index_folder must now skip it
    # (its stat was stored) and only index the never-seen notes.txt.
    outcome = _index_folder(str(docs_dir), db_path=db_path)
    assert outcome["files_skipped"] == 1
    assert outcome["files_indexed"] == 1


@pytest.mark.anyio
async def test_mcp_get_index_status_reports_db_size(mock_model, docs_dir, tmp_path):
    """get_index_status reports the index size on disk after indexing."""
    db_path = str(tmp_path / "testdb")
    from server.indexer import index_folder as _index_folder
    _index_folder(str(docs_dir), db_path=db_path)

    async with Client(mcp) as client:
        result = await client.call_tool("get_index_status", {"db_path": db_path})
        data = json.loads(result.content[0].text)

    assert "db_size_bytes" in data
    assert "db_size" in data
    assert data["db_size_bytes"] > 0  # a populated index occupies disk
    assert isinstance(data["db_size"], str)


@pytest.mark.anyio
async def test_mcp_get_index_status_db_size_zero_when_empty(tmp_path):
    """db_size is present and zero for a nonexistent index."""
    db_path = str(tmp_path / "emptydb")
    async with Client(mcp) as client:
        result = await client.call_tool("get_index_status", {"db_path": db_path})
        data = json.loads(result.content[0].text)

    assert data["db_size_bytes"] == 0
    assert data["db_size"] == "0 B"


@pytest.mark.anyio
async def test_mcp_reindex_file_skips_oversized(mock_model, tmp_path, monkeypatch):
    """reindex_file refuses a file over the size cap and leaves the index alone."""
    monkeypatch.setattr("server.indexer.MAX_FILE_SIZE_BYTES", 1024)
    big = tmp_path / "big.txt"
    big.write_text("x" * 5000)
    db_path = str(tmp_path / "db")

    async with Client(mcp) as client:
        result = await client.call_tool(
            "reindex_file", {"file_path": str(big), "db_path": db_path}
        )
        data = json.loads(result.content[0].text)

    assert data["status"] == "skipped"
    assert data["chunks_created"] == 0
    assert "cap" in data["reason"]


@pytest.mark.anyio
async def test_mcp_get_index_status_on_pre_tier2_index(tmp_path):
    """get_index_status (a read path) works on an un-migrated old-schema index."""
    import lancedb
    import pyarrow as pa
    from server.store import SCHEMA, TABLE_NAME

    db_path = str(tmp_path / "db")
    old_schema = pa.schema(
        [f for f in SCHEMA if f.name not in (
            "mtime_ns", "file_size", "chunk_kind", "sheet_name",
        )]
    )
    row = {
        "id": "x_0", "text": "hello", "source_file": "corpus/x.txt",
        "file_name": "x.txt", "file_type": ".txt", "folder_path": "corpus",
        "chunk_index": 0, "content_hash": "h", "vector": [0.0] * EMBEDDING_DIM,
    }
    lancedb.connect(db_path).create_table(TABLE_NAME, data=[row], schema=old_schema)

    async with Client(mcp) as client:
        result = await client.call_tool("get_index_status", {"db_path": db_path})
        data = json.loads(result.content[0].text)

    assert data["total_chunks"] == 1
    assert data["total_files"] == 1
    assert data["jobs"] == {"active": [], "recent": []}


# -- Exclusion rules: MCP surface --------------------------------------------

@pytest.mark.anyio
async def test_mcp_index_folder_accepts_exclude_param(mock_model, tmp_path):
    """The MCP tool accepts an `exclude` list and forwards it to indexing."""
    corpus = tmp_path / "corpus"
    corpus.mkdir()
    (corpus / "keep.md").write_text("real content")
    (corpus / "vendored").mkdir()
    (corpus / "vendored" / "junk.md").write_text("library README")

    db_path = str(tmp_path / "db")
    async with Client(mcp) as client:
        start = await client.call_tool(
            "index_folder",
            {
                "folder_path": str(corpus),
                "db_path": db_path,
                "exclude": ["vendored/**"],
            },
        )
        job_id = json.loads(start.content[0].text)["job_id"]
        job = await _wait_for_job(client, db_path, job_id)

    assert job["state"] == "completed"
    assert job["result"]["files_indexed"] == 1
    assert job["result"]["exclusion_patterns"] == ["vendored/**"]


@pytest.mark.anyio
async def test_mcp_index_folder_rejects_invalid_exclude_pattern(docs_dir, tmp_path):
    """A syntactically invalid exclusion pattern is rejected immediately —
    no job is started, no index touched, and the error message names the
    offending pattern."""
    db_path = str(tmp_path / "db")
    async with Client(mcp) as client:
        result = await client.call_tool(
            "index_folder",
            {
                "folder_path": str(docs_dir),
                "db_path": db_path,
                "exclude": ["\\"],  # bare backslash — invalid gitwildmatch
            },
        )
        data = json.loads(result.content[0].text)

    assert data["status"] == "rejected"
    assert data["reason"] == "invalid_exclusion_pattern"
    assert "\\" in data["message"]
    assert registry.has_running() is False  # no job was ever created


@pytest.mark.anyio
async def test_mcp_reindex_file_bypasses_semanticignore(mock_model, tmp_path):
    """reindex_file is an explicit per-file ask — it bypasses .semanticignore
    so a caller can still force-index a file that would otherwise be excluded."""
    corpus = tmp_path / "corpus"
    corpus.mkdir()
    (corpus / ".semanticignore").write_text("vendored/\n")
    (corpus / "vendored").mkdir()
    target = corpus / "vendored" / "force_me.md"
    target.write_text("explicitly requested content")

    db_path = str(tmp_path / "db")
    async with Client(mcp) as client:
        result = await client.call_tool(
            "reindex_file", {"file_path": str(target), "db_path": db_path}
        )
        data = json.loads(result.content[0].text)

    assert data["status"] == "reindexed"
    assert data["chunks_created"] >= 1


@pytest.mark.anyio
async def test_mcp_reindex_file_rejects_path_inside_index_dir(tmp_path):
    """reindex_file must refuse a path that lives inside the LanceDB directory
    — otherwise a stray call could try to parse the index's own files."""
    db_path = str(tmp_path / "db")
    db_dir = Path(db_path)
    db_dir.mkdir(parents=True)
    inside = db_dir / "data.lance" / "something.md"
    inside.parent.mkdir(parents=True)
    inside.write_text("# inside the index dir")

    async with Client(mcp) as client:
        result = await client.call_tool(
            "reindex_file", {"file_path": str(inside), "db_path": db_path}
        )
        data = json.loads(result.content[0].text)

    assert data["status"] == "rejected"
    assert data["reason"] == "inside_index_dir"


@pytest.mark.anyio
async def test_mcp_get_index_status_returns_exclusions_for_folder(mock_model, tmp_path):
    """When folder_path is supplied, get_index_status surfaces the active
    .semanticignore for that folder."""
    corpus = tmp_path / "corpus"
    corpus.mkdir()
    (corpus / ".semanticignore").write_text(
        "# header comment\n"
        "vendored/\n"
        "*.log\n"
    )
    db_path = str(tmp_path / "db")

    async with Client(mcp) as client:
        result = await client.call_tool(
            "get_index_status", {"db_path": db_path, "folder_path": str(corpus)}
        )
        data = json.loads(result.content[0].text)

    assert "exclusions" in data
    assert data["exclusions"] is not None
    assert data["exclusions"]["folder_path"] == str(corpus)
    assert data["exclusions"]["semanticignore_path"] == str(corpus / ".semanticignore")
    assert "vendored/" in data["exclusions"]["patterns"]
    assert "*.log" in data["exclusions"]["patterns"]


@pytest.mark.anyio
async def test_mcp_get_index_status_exclusions_null_when_no_file(mock_model, tmp_path):
    """When folder_path is supplied but no .semanticignore exists, exclusions
    is explicitly null — the caller can tell 'no rules' apart from 'not asked'."""
    corpus = tmp_path / "corpus"
    corpus.mkdir()
    db_path = str(tmp_path / "db")

    async with Client(mcp) as client:
        result = await client.call_tool(
            "get_index_status", {"db_path": db_path, "folder_path": str(corpus)}
        )
        data = json.loads(result.content[0].text)

    assert "exclusions" in data
    assert data["exclusions"] is None


@pytest.mark.anyio
async def test_mcp_get_index_status_omits_exclusions_without_folder_path(tmp_path):
    """When folder_path is NOT supplied, the exclusions field is omitted — the
    indexed root isn't persisted so there's nothing honest to report."""
    db_path = str(tmp_path / "db")
    async with Client(mcp) as client:
        result = await client.call_tool("get_index_status", {"db_path": db_path})
        data = json.loads(result.content[0].text)

    assert "exclusions" not in data


# --- spreadsheet description tools ----------------------------------------


@pytest.mark.anyio
async def test_list_pending_descriptions_returns_enqueued(tmp_path):
    from server.store import VectorStore
    from server.paths import to_relative

    db = str((tmp_path / "db").resolve())
    # File must exist on disk or list_pending_descriptions auto-evicts it.
    folder = tmp_path / "folder"
    folder.mkdir()
    xlsx = folder / "a.xlsx"
    xlsx.write_bytes(b"placeholder")  # contents don't matter for list_*
    rel = to_relative(str(xlsx), db)

    store = VectorStore(db)
    store.enqueue_pending(
        rel,
        ["sheet:Sales", "file"],
        json.dumps({"type": "xlsx", "sheets": [{"name": "Sales"}]}),
        "h1",
    )
    async with Client(mcp) as client:
        result = await client.call_tool(
            "list_pending_descriptions", {"db_path": db}
        )
    data = json.loads(result.content[0].text)
    assert data["total_remaining"] == 1
    assert len(data["pending"]) == 1
    entry = data["pending"][0]
    assert entry["file_path"] == rel
    assert entry["needs"] == ["sheet:Sales", "file"]
    assert entry["preview"]["type"] == "xlsx"


@pytest.mark.anyio
async def test_list_pending_descriptions_auto_evicts_deleted_files(tmp_path):
    from server.store import VectorStore

    db = str((tmp_path / "db").resolve())
    fake_rel = "missing/x.csv"  # file does not exist on disk
    store = VectorStore(db)
    store.enqueue_pending(fake_rel, ["file"], json.dumps({"type": "csv"}), "h")
    async with Client(mcp) as client:
        result = await client.call_tool(
            "list_pending_descriptions", {"db_path": db}
        )
    data = json.loads(result.content[0].text)
    assert data["pending"] == []
    # Auto-evicted from the table on the read path
    fresh = VectorStore(db)
    assert fresh.pending_count() == 0



@pytest.mark.anyio
async def test_submit_description_csv_completes_file(mock_model, tmp_path):
    from server.store import VectorStore
    from server.paths import to_relative

    folder = tmp_path / "data"
    folder.mkdir()
    csv = folder / "x.csv"
    csv.write_text("a,b\n1,2\n", encoding="utf-8")
    db = str((tmp_path / "db").resolve())
    rel = to_relative(str(csv), db)

    store = VectorStore(db)
    store.enqueue_pending(
        rel, ["file"],
        json.dumps({"type": "csv", "sheets": [{"name": "x.csv"}]}),
        "h",
    )

    async with Client(mcp) as client:
        result = await client.call_tool(
            "submit_description",
            {
                "file_path": str(csv),
                "sheet_name": None,
                "description": "This is a tiny customer-name list.",
                "db_path": db,
            },
        )
    data = json.loads(result.content[0].text)
    assert data["status"] == "stored"
    assert data["file_complete"] is True
    fresh = VectorStore(db)
    assert fresh.count_chunks() == 1
    assert fresh.pending_count() == 0


@pytest.mark.anyio
async def test_submit_description_xlsx_incremental(mock_model, tmp_path):
    import openpyxl
    from server.store import VectorStore
    from server.paths import to_relative

    folder = tmp_path / "data"
    folder.mkdir()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Sales").append(["a", "b"])
    wb.create_sheet("Costs").append(["c", "d"])
    xlsx = folder / "biz.xlsx"
    wb.save(xlsx)
    db = str((tmp_path / "db").resolve())
    rel = to_relative(str(xlsx), db)

    store = VectorStore(db)
    store.enqueue_pending(
        rel,
        ["sheet:Sales", "sheet:Costs", "file"],
        json.dumps({"type": "xlsx", "sheets": [
            {"name": "Sales"}, {"name": "Costs"},
        ]}),
        "h",
    )

    async with Client(mcp) as client:
        r1 = await client.call_tool("submit_description", {
            "file_path": str(xlsx), "sheet_name": "Sales",
            "description": "Sales by region", "db_path": db,
        })
        d1 = json.loads(r1.content[0].text)
        assert d1["file_complete"] is False

        r2 = await client.call_tool("submit_description", {
            "file_path": str(xlsx), "sheet_name": "Costs",
            "description": "Costs by category", "db_path": db,
        })
        d2 = json.loads(r2.content[0].text)
        assert d2["file_complete"] is False

        r3 = await client.call_tool("submit_description", {
            "file_path": str(xlsx), "sheet_name": None,
            "description": "Sales and costs ledger.", "db_path": db,
        })
        d3 = json.loads(r3.content[0].text)
        assert d3["file_complete"] is True

    fresh = VectorStore(db)
    assert fresh.count_chunks() == 3
    assert fresh.pending_count() == 0


@pytest.mark.anyio
async def test_submit_description_rejects_unknown_sheet(mock_model, tmp_path):
    from server.store import VectorStore
    from server.paths import to_relative

    folder = tmp_path / "data"
    folder.mkdir()
    csv = folder / "x.csv"
    csv.write_text("a,b\n1,2\n", encoding="utf-8")
    db = str((tmp_path / "db").resolve())
    rel = to_relative(str(csv), db)

    store = VectorStore(db)
    store.enqueue_pending(
        rel, ["file"], json.dumps({"type": "csv", "sheets": [{"name": "x.csv"}]}),
        "h",
    )

    async with Client(mcp) as client:
        result = await client.call_tool("submit_description", {
            "file_path": str(csv), "sheet_name": "NotInFile",
            "description": "x", "db_path": db,
        })
    data = json.loads(result.content[0].text)
    assert data["status"] == "rejected"
    fresh = VectorStore(db)
    assert fresh.pending_count() == 1


@pytest.mark.anyio
async def test_submit_description_rejects_empty(mock_model, tmp_path):
    from server.store import VectorStore
    from server.paths import to_relative

    folder = tmp_path / "data"
    folder.mkdir()
    csv = folder / "x.csv"
    csv.write_text("a,b\n1,2\n", encoding="utf-8")
    db = str((tmp_path / "db").resolve())
    rel = to_relative(str(csv), db)
    VectorStore(db).enqueue_pending(rel, ["file"], "{}", "h")

    async with Client(mcp) as client:
        result = await client.call_tool("submit_description", {
            "file_path": str(csv), "sheet_name": None,
            "description": "   ", "db_path": db,
        })
    data = json.loads(result.content[0].text)
    assert data["status"] == "rejected"
    assert data["reason"] == "empty_description"


@pytest.mark.anyio
async def test_dismiss_pending_description_removes_and_records_hash(tmp_path):
    from server.store import VectorStore
    from server.paths import to_relative
    from server.indexer import compute_file_hash

    folder = tmp_path / "data"
    folder.mkdir()
    csv = folder / "junk.csv"
    csv.write_text("a,b\n1,2\n", encoding="utf-8")
    db = str((tmp_path / "db").resolve())
    rel = to_relative(str(csv), db)
    h = compute_file_hash(csv)

    store = VectorStore(db)
    store.enqueue_pending(rel, ["file"], json.dumps({"type": "csv"}), h)

    async with Client(mcp) as client:
        result = await client.call_tool("dismiss_pending_description", {
            "file_path": str(csv), "db_path": db,
        })
    data = json.loads(result.content[0].text)
    assert data["status"] == "dismissed"
    fresh = VectorStore(db)
    assert fresh.pending_count() == 0
    assert fresh.is_dismissed(rel, h)


@pytest.mark.anyio
async def test_dismiss_pending_description_rejects_unknown(tmp_path):
    folder = tmp_path / "data"; folder.mkdir()
    csv = folder / "x.csv"
    csv.write_text("a,b\n1,2\n", encoding="utf-8")
    db = str((tmp_path / "db").resolve())

    async with Client(mcp) as client:
        result = await client.call_tool("dismiss_pending_description", {
            "file_path": str(csv), "db_path": db,
        })
    data = json.loads(result.content[0].text)
    assert data["status"] == "rejected"
    assert data["reason"] == "not_pending"


@pytest.mark.anyio
async def test_dismiss_pending_description_rejects_missing_file(tmp_path):
    db = str((tmp_path / "db").resolve())
    async with Client(mcp) as client:
        result = await client.call_tool("dismiss_pending_description", {
            "file_path": str(tmp_path / "nope.csv"), "db_path": db,
        })
    data = json.loads(result.content[0].text)
    assert data["status"] == "rejected"
    assert data["reason"] == "file_not_found"
