"""Tests for scripts/spreadsheets_to_txt_folder.py — standalone batch
spreadsheet (.csv/.xlsx/.xlsm/.xls) -> .txt converter (recursive, mirror-layout,
per-sheet, size-capped, run-log backed).

openpyxl is a project dependency so real workbooks are authored and converted
end-to-end; the unit under test is the script's discovery + mirror layout +
per-sheet split + size truncation + idempotency + error-isolation behaviour.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl


# -- helpers --------------------------------------------------------------


def _runlog_items(dst: Path, script: str = "spreadsheets_to_txt_folder") -> list[dict]:
    logs = sorted((dst / "_runlogs").glob(f"{script}-*.jsonl"))
    assert logs, "no run-log written"
    items = []
    for line in logs[-1].read_text(encoding="utf-8").splitlines():
        rec = json.loads(line)
        if rec["event"] == "item":
            items.append(rec)
    return items


def _write_csv(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _write_xlsx(path: Path, sheets: dict[str, list[list]]) -> None:
    """Author an xlsx. ``sheets`` maps sheet name -> list of rows (each a list
    of cell values); an empty row list creates an empty sheet."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default sheet; add exactly what's asked
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    wb.save(str(path))


# -- discovery + mirror layout -------------------------------------------


def test_csv_becomes_single_tab_separated_txt(tmp_path: Path) -> None:
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    _write_csv(src / "sub" / "data.csv", "a,b,c\n1,2,3\n4,5,6\n")
    _write_csv(src / "notes.md", "ignored\n")  # not a spreadsheet

    assert main([str(src), str(dst)]) == 0
    out = dst / "sub" / "data.txt"
    assert out.read_text(encoding="utf-8") == "a\tb\tc\n1\t2\t3\n4\t5\t6\n"
    assert not (dst / "notes.txt").exists()
    # Source untouched.
    assert (src / "sub" / "data.csv").is_file()


def test_workbook_splits_into_per_sheet_subdir(tmp_path: Path) -> None:
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    _write_xlsx(src / "report.xlsx", {
        "Sheet One": [["name", "val"], ["x", 10], ["y", 20.0]],
        "Second": [["h"], ["v"]],
    })

    assert main([str(src), str(dst)]) == 0
    one = dst / "report" / "Sheet One.txt"
    two = dst / "report" / "Second.txt"
    assert one.is_file() and two.is_file()
    # Integral floats normalise (20.0 -> "20"), tabs separate columns.
    assert one.read_text(encoding="utf-8") == "name\tval\nx\t10\ny\t20\n"


def test_legacy_xls_splits_into_per_sheet_subdir(tmp_path: Path) -> None:
    import pytest

    xlwt = pytest.importorskip("xlwt")

    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    src.mkdir(parents=True)
    wb = xlwt.Workbook()
    for sheet_name, rows in {
        "Sales": [["region", "qty"], ["EMEA", 10], ["APAC", 15]],
        "Notes": [["note"], ["ok"]],
    }.items():
        ws = wb.add_sheet(sheet_name)
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                ws.write(r, c, val)
    wb.save(str(src / "legacy.xls"))

    assert main([str(src), str(dst)]) == 0
    sales = dst / "legacy" / "Sales.txt"
    assert sales.is_file()
    # xlrd surfaces numbers as floats; integral floats normalise (10 not 10.0).
    assert sales.read_text(encoding="utf-8") == "region\tqty\nEMEA\t10\nAPAC\t15\n"
    assert (dst / "legacy" / "Notes.txt").is_file()


def test_full_extension_set_case_insensitive(tmp_path: Path) -> None:
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    _write_csv(src / "a.CSV", "x\n1\n")
    _write_xlsx(src / "b.XLSX", {"S": [["h"], ["v"]]})
    _write_xlsx(src / "c.xlsm", {"S": [["h"], ["v"]]})
    _write_csv(src / "skip.txt", "nope\n")  # not a recognised spreadsheet

    assert main([str(src), str(dst)]) == 0
    assert (dst / "a.txt").is_file()
    assert (dst / "b" / "S.txt").is_file()
    assert (dst / "c" / "S.txt").is_file()
    assert not (dst / "skip.txt").exists()


def test_empty_sheets_are_skipped(tmp_path: Path) -> None:
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    _write_xlsx(src / "wb.xlsx", {"Empty": [], "Data": [["h"], ["v"]]})

    assert main([str(src), str(dst)]) == 0
    assert not (dst / "wb" / "Empty.txt").exists()
    assert (dst / "wb" / "Data.txt").is_file()


# -- size truncation ------------------------------------------------------


def test_large_sheet_is_truncated_with_marker_under_cap(tmp_path: Path) -> None:
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    rows: list[list] = [["col1", "col2"]]
    rows += [[f"value-{i}", i] for i in range(5000)]
    _write_xlsx(src / "big.xlsx", {"Big": rows})

    # 0.01 MB target -> definitely truncates the 5001-row sheet.
    assert main([str(src), str(dst), "--target-mb", "0.01"]) == 0
    out = dst / "big" / "Big.txt"
    lines = out.read_text(encoding="utf-8").splitlines()

    assert lines[-1].startswith("... [truncated: wrote ")
    assert "of 5001 rows" in lines[-1]
    # Fewer data lines than the source, and well under the 20 MB hard cap.
    assert len(lines) - 1 < 5001
    assert out.stat().st_size <= 20 * 1024 * 1024


def test_max_mb_is_a_hard_cap_even_for_opening_rows(tmp_path: Path) -> None:
    """The header + first data row are kept past the soft target, but never past
    the hard --max-mb cap, so the output is guaranteed under it."""
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    # Two ~5 KB cells: the header+first-row grace would exceed a 4 KB cap.
    big_cell = "z" * 5000
    _write_xlsx(src / "wb.xlsx", {"S": [[big_cell], [big_cell], ["tiny"]]})

    cap_bytes = 4 * 1024
    assert main([
        str(src), str(dst), "--target-mb", "0.001", "--max-mb", "0.004",
    ]) == 0
    out = dst / "wb" / "S.txt"
    assert out.stat().st_size <= cap_bytes  # hard cap honoured


def test_large_csv_field_does_not_fail(tmp_path: Path) -> None:
    """A CSV cell larger than csv's default 128 KB field limit converts fine."""
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    big = "x" * 200_000  # > 131072 default field-size limit
    _write_csv(src / "data.csv", f"col\n{big}\n")

    assert main([str(src), str(dst)]) == 0
    out = dst / "data.txt"
    assert out.is_file()
    assert big in out.read_text(encoding="utf-8")


def test_target_must_be_below_max(tmp_path: Path) -> None:
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    _write_csv(src / "data.csv", "a\n1\n")

    rc = main([str(src), str(dst), "--target-mb", "20", "--max-mb", "20"])
    assert rc == 2
    assert not dst.exists() or not (dst / "data.txt").exists()


# -- idempotent re-run ----------------------------------------------------


def test_rerun_skips_done_and_writes_nothing_new(tmp_path: Path) -> None:
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    _write_csv(src / "data.csv", "a,b\n1,2\n")
    _write_xlsx(src / "wb.xlsx", {"S": [["h"], ["v"]]})

    assert main([str(src), str(dst)]) == 0
    snapshot = {p: p.stat().st_mtime_ns for p in dst.rglob("*.txt")}

    assert main([str(src), str(dst)]) == 0
    assert {it["status"] for it in _runlog_items(dst)} == {"skip"}
    assert {p: p.stat().st_mtime_ns for p in dst.rglob("*.txt")} == snapshot


def test_rerun_skips_workbook_with_empty_first_sheet(tmp_path: Path) -> None:
    """Regression: done-ness keys on the workbook's subdir, not a first-sheet
    txt that may not exist when sheet 1 is empty."""
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    _write_xlsx(src / "wb.xlsx", {"Empty": [], "Data": [["h"], ["v"]]})

    assert main([str(src), str(dst)]) == 0
    assert not (dst / "wb" / "Empty.txt").exists()  # first sheet produced no file

    assert main([str(src), str(dst)]) == 0
    statuses = {it["input"]: it["status"] for it in _runlog_items(dst)}
    assert statuses == {"wb.xlsx": "skip"}


def test_rerun_skips_all_empty_workbook(tmp_path: Path) -> None:
    """An all-empty workbook still creates its subdir, records as done, and is
    skipped on re-run (rather than reprocessed forever)."""
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    _write_xlsx(src / "blank.xlsx", {"Empty": []})

    assert main([str(src), str(dst)]) == 0
    assert (dst / "blank").is_dir()
    assert not list((dst / "blank").glob("*.txt"))

    assert main([str(src), str(dst)]) == 0
    statuses = {it["input"]: it["status"] for it in _runlog_items(dst)}
    assert statuses == {"blank.xlsx": "skip"}


def test_force_rerun_reuses_subdir_without_duplicates(tmp_path: Path) -> None:
    """Regression: --force clears and rewrites the same subdir; no <sheet>-1.txt."""
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    _write_xlsx(src / "wb.xlsx", {"Alpha": [["h"], ["v"]], "Beta": [["h"], ["v"]]})

    assert main([str(src), str(dst)]) == 0
    before = sorted(p.name for p in (dst / "wb").glob("*.txt"))

    assert main([str(src), str(dst), "--force"]) == 0
    after = sorted(p.name for p in (dst / "wb").glob("*.txt"))

    assert before == after == ["Alpha.txt", "Beta.txt"]
    assert not list((dst / "wb").glob("*-1.txt"))


# -- per-file error isolation --------------------------------------------


def test_one_failure_does_not_abort_batch(tmp_path: Path, capsys) -> None:
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    _write_csv(src / "good.csv", "a\n1\n")
    # A .xlsx that is not a valid zip -> UnreadableSpreadsheetError.
    (src / "bad.xlsx").write_bytes(b"not a real xlsx")

    rc = main([str(src), str(dst)])
    assert rc == 1  # >=1 failure -> non-zero

    assert (dst / "good.txt").is_file()
    assert not (dst / "bad").exists()
    err = capsys.readouterr().err
    assert "bad.xlsx" in err
    statuses = {it["input"]: it["status"] for it in _runlog_items(dst)}
    assert statuses == {"good.csv": "ok", "bad.xlsx": "fail"}


def test_retry_after_fix_only_reprocesses_the_failure(tmp_path: Path) -> None:
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    _write_csv(src / "good.csv", "a\n1\n")
    (src / "bad.xlsx").write_bytes(b"not a real xlsx")

    assert main([str(src), str(dst)]) == 1

    # "Fix" the bad file by replacing it with a valid workbook, then re-run.
    _write_xlsx(src / "bad.xlsx", {"S": [["h"], ["v"]]})
    assert main([str(src), str(dst)]) == 0

    statuses = {it["input"]: it["status"] for it in _runlog_items(dst)}
    assert statuses == {"good.csv": "skip", "bad.xlsx": "ok"}


# -- guards + dry run -----------------------------------------------------


def test_input_must_be_existing_directory(tmp_path: Path) -> None:
    from scripts.spreadsheets_to_txt_folder import main

    assert main([str(tmp_path / "nope"), str(tmp_path / "out")]) != 0


def test_output_inside_input_refused(tmp_path: Path) -> None:
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    src.mkdir()
    _write_csv(src / "data.csv", "a\n1\n")
    inside = src / "out"

    assert main([str(src), str(inside)]) != 0
    assert not inside.exists() or not any(inside.rglob("*.txt"))


def test_dry_run_lists_targets_and_writes_nothing(tmp_path: Path, capsys) -> None:
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    _write_csv(src / "data.csv", "a\n1\n")
    _write_xlsx(src / "wb.xlsx", {"S": [["h"], ["v"]], "T": [["h"], ["v"]]})

    assert main([str(src), str(dst), "--dry-run"]) == 0
    assert not list(dst.rglob("*.txt"))
    assert not (dst / "_runlogs").exists()
    out = capsys.readouterr().err
    assert "data.csv" in out and "wb.xlsx" in out
    assert "2 sheet(s)" in out


def test_summary_line_reports_counts(tmp_path: Path, capsys) -> None:
    from scripts.spreadsheets_to_txt_folder import main

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    _write_csv(src / "data.csv", "a\n1\n")

    assert main([str(src), str(dst)]) == 0
    assert "Converted 1" in capsys.readouterr().err
