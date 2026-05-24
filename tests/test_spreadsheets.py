"""Unit tests for spreadsheet preview extraction and prompt building."""

from pathlib import Path

import pytest

from server.spreadsheets import _extract_csv_preview


def _write(tmp_path: Path, name: str, content: str) -> Path:
    p = tmp_path / name
    p.write_text(content, encoding="utf-8")
    return p


def test_csv_preview_with_header(tmp_path: Path):
    csv = _write(
        tmp_path, "customers.csv",
        "id,name,amount\n1,Alice,10.5\n2,Bob,20\n3,Carol,33.7\n",
    )
    sheet = _extract_csv_preview(csv)
    assert sheet["name"] == "customers.csv"
    assert sheet["headers"] == ["id", "name", "amount"]
    assert sheet["first_row_as_data"] is False
    assert sheet["dtypes"] == ["number", "string", "number"]
    assert sheet["row_count"] == 3
    assert sheet["sample_rows"][0] == ["1", "Alice", "10.5"]
    assert len(sheet["sample_rows"]) == 3


def test_csv_preview_no_header_sniffer_falls_back(tmp_path: Path):
    # All-numeric rows: Sniffer.has_header returns False
    csv = _write(tmp_path, "nums.csv", "1,2,3\n4,5,6\n7,8,9\n")
    sheet = _extract_csv_preview(csv)
    assert sheet["headers"] is None
    assert sheet["first_row_as_data"] is True
    assert sheet["row_count"] == 3


def test_csv_preview_empty_file(tmp_path: Path):
    csv = _write(tmp_path, "empty.csv", "")
    sheet = _extract_csv_preview(csv)
    assert sheet["headers"] is None
    assert sheet["row_count"] == 0
    assert sheet["sample_rows"] == []


def test_csv_preview_caps_sample_rows(tmp_path: Path):
    # Multi-column with a clear string header so Sniffer reliably detects it.
    rows = "name,value\n" + "\n".join(f"item{i},{i}" for i in range(100)) + "\n"
    csv = _write(tmp_path, "many.csv", rows)
    sheet = _extract_csv_preview(csv)
    assert len(sheet["sample_rows"]) == 10
    assert sheet["row_count"] == 100


def test_csv_preview_streams_no_whole_file_read(tmp_path: Path, monkeypatch):
    """Regression: _extract_csv_preview must stream — never call read_text,
    which loads the whole file into memory."""
    import pathlib
    rows = "name,value\n" + "\n".join(f"item{i},{i}" for i in range(100)) + "\n"
    csv = _write(tmp_path, "many.csv", rows)

    def boom(self, *args, **kwargs):
        raise AssertionError("read_text() called — CSV preview must stream")

    monkeypatch.setattr(pathlib.Path, "read_text", boom)
    sheet = _extract_csv_preview(csv)
    assert sheet["row_count"] == 100
    assert len(sheet["sample_rows"]) == 10
    assert sheet["headers"] == ["name", "value"]


def test_csv_preview_large_row_count_accurate(tmp_path: Path):
    """50K rows: streaming preserves exact row_count without materialising
    the whole file."""
    rows = ["id,val\n"]
    rows.extend(f"{i},x{i}\n" for i in range(50_000))
    csv = _write(tmp_path, "big.csv", "".join(rows))
    sheet = _extract_csv_preview(csv)
    assert sheet["row_count"] == 50_000
    assert sheet["headers"] == ["id", "val"]
    assert len(sheet["sample_rows"]) == 10
    # First data row is row 0 of the body, not the header.
    assert sheet["sample_rows"][0] == ["0", "x0"]


# --- XLSX preview ----------------------------------------------------------

import openpyxl

from server.spreadsheets import (
    _extract_xlsx_preview,
    UnreadableSpreadsheetError,
)


def _build_xlsx(tmp_path: Path, name: str, sheets: dict[str, list[list]]) -> Path:
    wb = openpyxl.Workbook()
    # remove default Sheet so we control the sheet set fully
    default = wb.active
    wb.remove(default)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    p = tmp_path / name
    wb.save(p)
    return p


def test_xlsx_preview_multi_sheet(tmp_path: Path):
    xlsx = _build_xlsx(tmp_path, "biz.xlsx", {
        "Sales": [
            ["region", "qty", "revenue"],
            ["EMEA", 10, 1000.0],
            ["APAC", 15, 1500.0],
        ],
        "Costs": [
            ["category", "amount"],
            ["rent", 5000],
            ["salaries", 25000],
        ],
    })
    sheets = _extract_xlsx_preview(xlsx)
    assert [s["name"] for s in sheets] == ["Sales", "Costs"]
    sales = sheets[0]
    assert sales["headers"] == ["region", "qty", "revenue"]
    assert sales["dtypes"] == ["string", "number", "number"]
    assert sales["row_count"] == 2
    # openpyxl's data_only normalises 1000.0 → 1000 when it fits an int
    assert sales["sample_rows"][0] == ["EMEA", "10", "1000"]


def test_xlsx_preview_header_only_sheet(tmp_path: Path):
    xlsx = _build_xlsx(tmp_path, "headers_only.xlsx", {
        "Empty": [["a", "b", "c"]],
    })
    sheets = _extract_xlsx_preview(xlsx)
    assert len(sheets) == 1
    assert sheets[0]["headers"] == ["a", "b", "c"]
    assert sheets[0]["row_count"] == 0
    assert sheets[0]["sample_rows"] == []


def test_xlsx_preview_empty_sheet(tmp_path: Path):
    xlsx = _build_xlsx(tmp_path, "empty_sheet.xlsx", {"Blank": []})
    sheets = _extract_xlsx_preview(xlsx)
    assert sheets[0]["headers"] is None
    assert sheets[0]["row_count"] == 0


def test_xlsx_preview_caps_sample_rows(tmp_path: Path):
    rows = [["h"]] + [[i] for i in range(100)]
    xlsx = _build_xlsx(tmp_path, "many.xlsx", {"S": rows})
    sheets = _extract_xlsx_preview(xlsx)
    assert len(sheets[0]["sample_rows"]) == 10
    assert sheets[0]["row_count"] == 100


def test_xlsx_preview_unreadable_file(tmp_path: Path):
    bad = tmp_path / "bogus.xlsx"
    bad.write_bytes(b"not actually a zip")
    with pytest.raises(UnreadableSpreadsheetError):
        _extract_xlsx_preview(bad)


# --- XLSM (macro-enabled xlsx) ---------------------------------------------
#
# Same OOXML container as .xlsx — openpyxl reads it identically. We build via
# the .xlsx writer and rename, so we don't have to authenticate a fake VBA
# archive to make openpyxl save with the .xlsm content type.


def _build_xlsm(tmp_path: Path, name: str, sheets: dict[str, list[list]]) -> Path:
    tmp = _build_xlsx(tmp_path, "_tmp.xlsx", sheets)
    target = tmp_path / name
    tmp.rename(target)
    return target


def test_xlsm_routes_through_xlsx_preview(tmp_path: Path):
    xlsm = _build_xlsm(tmp_path, "wb.xlsm", {
        "Sales": [["region", "qty"], ["EMEA", 10]],
        "Costs": [["category", "amount"], ["rent", 5000]],
    })
    sheets = _extract_xlsx_preview(xlsm)
    assert [s["name"] for s in sheets] == ["Sales", "Costs"]
    assert sheets[0]["headers"] == ["region", "qty"]


# --- XLS (legacy BIFF) -----------------------------------------------------


def _build_xls(folder: Path, name: str, sheets: dict[str, list[list]]) -> Path:
    import xlwt
    wb = xlwt.Workbook()
    for sheet_name, rows in sheets.items():
        ws = wb.add_sheet(sheet_name)
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                ws.write(r, c, val)
    p = folder / name
    wb.save(str(p))
    return p


def test_xls_preview_basic(tmp_path: Path):
    from server.spreadsheets import _extract_xls_preview
    xls = _build_xls(tmp_path, "legacy.xls", {
        "Sales": [["region", "qty"], ["EMEA", 10], ["APAC", 15]],
    })
    sheets = _extract_xls_preview(xls)
    assert len(sheets) == 1
    s = sheets[0]
    assert s["name"] == "Sales"
    assert s["headers"] == ["region", "qty"]
    assert s["dtypes"] == ["string", "number"]
    assert s["row_count"] == 2
    # xlrd surfaces all numbers as floats; preview must normalise integral
    # values so "10" not "10.0" lands in the LLM prompt.
    assert s["sample_rows"][0] == ["EMEA", "10"]


def test_xls_preview_multi_sheet(tmp_path: Path):
    from server.spreadsheets import _extract_xls_preview
    xls = _build_xls(tmp_path, "biz.xls", {
        "Sales": [["a", "b"], [1, 2]],
        "Costs": [["c", "d"], [3, 4]],
    })
    sheets = _extract_xls_preview(xls)
    assert [s["name"] for s in sheets] == ["Sales", "Costs"]


def test_xls_preview_unreadable_file(tmp_path: Path):
    from server.spreadsheets import _extract_xls_preview
    bad = tmp_path / "bogus.xls"
    bad.write_bytes(b"definitely not a BIFF compound document")
    with pytest.raises(UnreadableSpreadsheetError):
        _extract_xls_preview(bad)


# --- Dispatcher + prompt builders ------------------------------------------

from server.spreadsheets import (
    extract_preview,
    build_sheet_prompt,
    build_file_prompt,
    needs_for_preview,
    SPREADSHEET_EXTENSIONS,
)


def test_extract_preview_csv(tmp_path: Path):
    csv = _write(tmp_path, "x.csv", "a,b\n1,2\n3,4\n")
    pv = extract_preview(csv)
    assert pv["type"] == "csv"
    assert len(pv["sheets"]) == 1
    assert pv["sheets"][0]["headers"] == ["a", "b"]


def test_extract_preview_xlsx(tmp_path: Path):
    xlsx = _build_xlsx(tmp_path, "x.xlsx", {"S": [["a", "b"], [1, 2]]})
    pv = extract_preview(xlsx)
    assert pv["type"] == "xlsx"
    assert pv["sheets"][0]["name"] == "S"


def test_extract_preview_rejects_unknown(tmp_path: Path):
    p = _write(tmp_path, "x.txt", "hello")
    with pytest.raises(ValueError):
        extract_preview(p)


def test_supported_extensions_constant():
    assert ".csv" in SPREADSHEET_EXTENSIONS
    assert ".xlsx" in SPREADSHEET_EXTENSIONS
    assert ".xlsm" in SPREADSHEET_EXTENSIONS
    assert ".xls" in SPREADSHEET_EXTENSIONS


def test_extract_preview_xlsm(tmp_path: Path):
    xlsm = _build_xlsm(tmp_path, "x.xlsm", {"S": [["a", "b"], [1, 2]]})
    pv = extract_preview(xlsm)
    assert pv["type"] == "xlsm"
    assert pv["sheets"][0]["name"] == "S"


def test_extract_preview_xls(tmp_path: Path):
    xls = _build_xls(tmp_path, "x.xls", {"S": [["a", "b"], [1, 2]]})
    pv = extract_preview(xls)
    assert pv["type"] == "xls"
    assert pv["sheets"][0]["name"] == "S"


def test_needs_for_xls_and_xlsm_match_workbook_path():
    """Non-csv types fall through to the workbook needs path."""
    for t in ("xls", "xlsm"):
        pv = {"type": t, "sheets": [{"name": "S1"}, {"name": "S2"}]}
        assert needs_for_preview(pv) == ["sheet:S1", "sheet:S2", "file"]


def test_needs_for_csv():
    pv = {"type": "csv", "sheets": [{"name": "x.csv"}]}
    assert needs_for_preview(pv) == ["file"]


def test_needs_for_xlsx():
    pv = {"type": "xlsx", "sheets": [{"name": "Sales"}, {"name": "Costs"}]}
    assert needs_for_preview(pv) == ["sheet:Sales", "sheet:Costs", "file"]


def test_build_sheet_prompt_includes_context():
    pv = {"type": "xlsx", "sheets": [{
        "name": "Sales",
        "headers": ["region", "qty"],
        "dtypes": ["string", "number"],
        "row_count": 1000,
        "sample_rows": [["EMEA", "10"]],
        "first_row_as_data": False,
    }]}
    prompt = build_sheet_prompt(pv, "Sales")
    assert "Sales" in prompt
    assert "region" in prompt
    assert "qty" in prompt
    assert "1000" in prompt
    assert "EMEA" in prompt


def test_build_file_prompt_includes_sheet_descriptions():
    pv = {"type": "xlsx", "sheets": [
        {"name": "Sales", "headers": ["a"], "dtypes": ["string"],
         "row_count": 1, "sample_rows": [], "first_row_as_data": False},
        {"name": "Costs", "headers": ["b"], "dtypes": ["string"],
         "row_count": 1, "sample_rows": [], "first_row_as_data": False},
    ]}
    descs = {"Sales": "Tracks sales by region.", "Costs": "Lists expenses."}
    prompt = build_file_prompt(pv, descs)
    assert "Tracks sales by region." in prompt
    assert "Lists expenses." in prompt


def test_build_file_prompt_csv_uses_single_sheet_block():
    pv = {"type": "csv", "sheets": [{
        "name": "data.csv",
        "headers": ["id", "name"],
        "dtypes": ["number", "string"],
        "row_count": 42,
        "sample_rows": [["1", "alice"]],
        "first_row_as_data": False,
    }]}
    prompt = build_file_prompt(pv, {})
    assert "data.csv" in prompt
    assert "42" in prompt
    assert "alice" in prompt
