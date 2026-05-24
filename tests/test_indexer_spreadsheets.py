"""Integration tests for the spreadsheet → queue indexing path.

Spreadsheet indexing is *temporarily disabled* (see
server/parsers.py:SUPPORTED_EXTENSIONS). The legacy enqueue/auto-drain
tests below are wrapped in `requires_spreadsheets_enabled` so they're
skipped as a unit when the disable is in place — un-decorate them when
spreadsheets are re-enabled. New tests at the bottom of this file assert
the disabled behaviour (filtered from discovery, existing chunks
preserved).
"""

import json
from pathlib import Path

import openpyxl
import pytest

from server.indexer import index_folder, compute_file_hash
from server.parsers import SUPPORTED_EXTENSIONS
from server.paths import to_relative
from server.store import VectorStore


# Wrap every legacy test that depends on spreadsheets being discovered. The
# block of tests below all assume index_folder will enqueue/sample a CSV or
# workbook; that path is unreachable while spreadsheets are disabled. The
# skip is keyed off the current SUPPORTED_EXTENSIONS so it lifts itself
# automatically once the disable is removed.
requires_spreadsheets_enabled = pytest.mark.skipif(
    not (SUPPORTED_EXTENSIONS & {".csv", ".xlsx", ".xlsm", ".xls"}),
    reason="spreadsheet indexing temporarily disabled "
           "(see server/parsers.py:SUPPORTED_EXTENSIONS)",
)


def _write_csv(folder: Path, name: str) -> Path:
    p = folder / name
    p.write_text("id,name\n1,Alice\n2,Bob\n", encoding="utf-8")
    return p


def _write_xlsx(folder: Path, name: str, sheets: dict[str, list[list]]) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    p = folder / name
    wb.save(p)
    return p


@requires_spreadsheets_enabled
def test_index_folder_enqueues_csv(tmp_path):
    folder = tmp_path / "data"
    folder.mkdir()
    _write_csv(folder, "x.csv")
    db = str((tmp_path / "db").resolve())

    result = index_folder(str(folder), db_path=db)

    assert result["descriptions_queued"] == 1
    assert result["descriptions_sampled"] == 0
    store = VectorStore(db)
    assert store.count_chunks() == 0
    pending = store.list_pending()
    assert len(pending) == 1
    assert pending[0]["needs"] == ["file"]
    preview = json.loads(pending[0]["preview_json"])
    assert preview["type"] == "csv"


@requires_spreadsheets_enabled
def test_index_folder_enqueues_xlsx_per_sheet(tmp_path):
    folder = tmp_path / "data"
    folder.mkdir()
    _write_xlsx(folder, "biz.xlsx", {
        "Sales": [["a", "b"], [1, 2]],
        "Costs": [["c", "d"], [3, 4]],
    })
    db = str((tmp_path / "db").resolve())

    result = index_folder(str(folder), db_path=db)

    assert result["descriptions_queued"] == 1  # one file (with 3 needs)
    store = VectorStore(db)
    [pending] = store.list_pending()
    assert pending["needs"] == ["sheet:Sales", "sheet:Costs", "file"]


def _write_xlsm(folder: Path, name: str, sheets: dict[str, list[list]]) -> Path:
    """An .xlsm is the same OOXML container as .xlsx — write then rename so
    openpyxl's writer doesn't object to a fresh workbook lacking VBA."""
    tmp = _write_xlsx(folder, "_tmp.xlsx", sheets)
    target = folder / name
    tmp.rename(target)
    return target


def _write_xls(folder: Path, name: str, sheets: dict[str, list[list]]) -> Path:
    import xlwt
    wb = xlwt.Workbook()
    for sheet_name, rows in sheets.items():
        ws = wb.add_sheet(sheet_name)
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                ws.write(r, c, val)
    p = folder / name
    wb.save(str(p))
    return p


@requires_spreadsheets_enabled
def test_index_folder_enqueues_xlsm_per_sheet(tmp_path):
    folder = tmp_path / "data"
    folder.mkdir()
    _write_xlsm(folder, "macros.xlsm", {
        "Sales": [["a", "b"], [1, 2]],
        "Costs": [["c", "d"], [3, 4]],
    })
    db = str((tmp_path / "db").resolve())

    result = index_folder(str(folder), db_path=db)

    assert result["descriptions_queued"] == 1
    store = VectorStore(db)
    [pending] = store.list_pending()
    assert pending["needs"] == ["sheet:Sales", "sheet:Costs", "file"]
    assert json.loads(pending["preview_json"])["type"] == "xlsm"


@requires_spreadsheets_enabled
def test_index_folder_enqueues_xls_per_sheet(tmp_path):
    folder = tmp_path / "data"
    folder.mkdir()
    _write_xls(folder, "legacy.xls", {
        "Sales": [["a", "b"], [1, 2]],
        "Costs": [["c", "d"], [3, 4]],
    })
    db = str((tmp_path / "db").resolve())

    result = index_folder(str(folder), db_path=db)

    assert result["descriptions_queued"] == 1
    store = VectorStore(db)
    [pending] = store.list_pending()
    assert pending["needs"] == ["sheet:Sales", "sheet:Costs", "file"]
    assert json.loads(pending["preview_json"])["type"] == "xls"


@requires_spreadsheets_enabled
def test_index_folder_skips_dismissed_with_matching_hash(tmp_path):
    folder = tmp_path / "data"
    folder.mkdir()
    csv = _write_csv(folder, "x.csv")
    db = str((tmp_path / "db").resolve())

    h = compute_file_hash(csv)
    rel = to_relative(str(csv), db)
    store = VectorStore(db)
    store.dismiss(rel, h)

    result = index_folder(str(folder), db_path=db)
    assert result["descriptions_queued"] == 0
    assert store.pending_count() == 0


@requires_spreadsheets_enabled
def test_index_folder_reenqueues_when_dismissed_hash_differs(tmp_path):
    folder = tmp_path / "data"
    folder.mkdir()
    csv = _write_csv(folder, "x.csv")
    db = str((tmp_path / "db").resolve())
    rel = to_relative(str(csv), db)
    store = VectorStore(db)
    store.dismiss(rel, "stale_hash_from_old_content")

    result = index_folder(str(folder), db_path=db)
    assert result["descriptions_queued"] == 1


def test_exceeds_size_cap_exempts_spreadsheets(monkeypatch):
    """Spreadsheet parsers stream their preview, so the global 100 MB cap
    must not skip them regardless of file size."""
    monkeypatch.setattr("server.indexer.MAX_FILE_SIZE_BYTES", 100)
    from server.indexer import exceeds_size_cap
    for suffix in (".csv", ".xlsx", ".xlsm", ".xls"):
        assert exceeds_size_cap(Path(f"x{suffix}"), 1_000_000_000) is False, suffix


def test_exceeds_size_cap_still_applies_to_non_streaming(monkeypatch):
    """Sanity: PDFs (which load the whole file) still respect the cap."""
    monkeypatch.setattr("server.indexer.MAX_FILE_SIZE_BYTES", 100)
    from server.indexer import exceeds_size_cap
    assert exceeds_size_cap(Path("x.pdf"), 200) is True


@requires_spreadsheets_enabled
def test_index_folder_streams_large_csv_above_size_cap(tmp_path, monkeypatch):
    """A CSV bigger than the cap is enqueued via streaming, not size_skipped."""
    monkeypatch.setattr("server.indexer.MAX_FILE_SIZE_BYTES", 100)
    folder = tmp_path / "data"
    folder.mkdir()
    rows = "id,val\n" + "\n".join(f"{i},x{i}" for i in range(200))
    (folder / "big.csv").write_text(rows, encoding="utf-8")
    db = str((tmp_path / "db").resolve())

    result = index_folder(str(folder), db_path=db)

    assert result["descriptions_queued"] == 1
    assert result["files_size_skipped"] == 0
    assert result["oversized_files"] == []


@requires_spreadsheets_enabled
def test_index_folder_evicts_legacy_csv_chunks_then_enqueues(tmp_path):
    """A pre-existing CSV indexed under the legacy raw-row scheme is evicted
    on the first run and replaced with a queue entry."""
    folder = tmp_path / "data"
    folder.mkdir()
    csv = _write_csv(folder, "x.csv")
    db = str((tmp_path / "db").resolve())
    rel = to_relative(str(csv), db)

    # Seed the store with a legacy "raw row text" chunk for this CSV.
    import numpy as np
    from server.store import EMBEDDING_DIM
    vec = np.random.RandomState(0).randn(EMBEDDING_DIM).astype(np.float32)
    vec = (vec / np.linalg.norm(vec)).tolist()
    store = VectorStore(db)
    store.add_chunks([{
        "id": "legacy_0",
        "text": "id,name\n1,Alice\n2,Bob",
        "source_file": rel,
        "file_name": "x.csv",
        "file_type": ".csv",
        "folder_path": str(folder.relative_to(folder.parent)),
        "chunk_index": 0,
        "content_hash": "ignored",
        "mtime_ns": 0,
        "file_size": 0,
        "vector": vec,
        # chunk_kind intentionally omitted → defaults to "text" via add_chunks
    }])
    assert store.count_chunks() == 1

    result = index_folder(str(folder), db_path=db)
    # Legacy chunk gone, CSV re-routed through the queue. Fresh handle —
    # LanceDB's cached table handle on `store` won't see writes from the
    # connection that `index_folder` opened internally.
    fresh = VectorStore(db)
    assert fresh.count_chunks() == 0
    assert result["descriptions_queued"] == 1
    assert fresh.pending_count() == 1


# --- auto-drainer (FastMCP sampling) ---


import asyncio
from unittest.mock import AsyncMock, MagicMock, patch

import numpy as np


def _fake_embed(texts):
    out = []
    for t in texts:
        rng = np.random.RandomState(hash(t) % 2**32)
        v = rng.randn(256).astype("float32")
        out.append(v / np.linalg.norm(v))
    return np.array(out)


@pytest.fixture
def mock_embed_model():
    model = type("M", (), {
        "encode": lambda self, texts, **kw: _fake_embed(texts)
    })()
    with patch("server.indexer.get_model", return_value=model):
        yield model


@requires_spreadsheets_enabled
def test_auto_drainer_writes_chunks_via_mock_sampling(mock_embed_model, tmp_path):
    """When ctx.sample is available, the auto-drainer turns queued entries
    into description chunks and clears the queue."""
    from server.main import _run_index_job
    from server.jobs import registry

    registry._jobs.clear(); registry._order.clear()
    registry._persist_path = None

    folder = tmp_path / "data"
    folder.mkdir()
    _write_xlsx(folder, "biz.xlsx", {
        "Sales": [["region", "qty"], ["EMEA", 10]],
        "Costs": [["category", "amount"], ["rent", 5000]],
    })
    db = str((tmp_path / "db").resolve())

    canned_texts = [
        "Sales by region.",
        "Cost breakdown by category.",
        "Workbook tracking sales vs costs.",
    ]
    canned_iter = iter(canned_texts)

    async def fake_sample(*args, **kwargs):
        return MagicMock(text=next(canned_iter))

    ctx = MagicMock()
    ctx.sample = fake_sample

    job = registry.create(str(folder), db)
    asyncio.run(_run_index_job(job, None, True, None, ctx=ctx))

    fresh = VectorStore(db)
    assert fresh.count_chunks() == 3
    assert fresh.pending_count() == 0
    final = job.to_dict()
    assert final["result"]["descriptions_sampled"] == 3


@requires_spreadsheets_enabled
def test_auto_drainer_leaves_entry_when_sampling_fails(
    mock_embed_model, tmp_path, monkeypatch,
):
    """A failure inside a file aborts that file atomically: no chunks land
    and the queue entry stays for manual drain later."""
    # Disable the 2s sampling-retry backoff so the test is fast.
    monkeypatch.setattr("server.main._SAMPLE_RETRY_BACKOFF_SECONDS", 0)
    from server.main import _run_index_job
    from server.jobs import registry

    registry._jobs.clear(); registry._order.clear()
    registry._persist_path = None

    folder = tmp_path / "data"
    folder.mkdir()
    _write_xlsx(folder, "biz.xlsx", {
        "Sales": [["a"], [1]],
        "Costs": [["b"], [2]],
    })
    db = str((tmp_path / "db").resolve())

    call_n = {"i": 0}
    async def flaky_sample(*args, **kwargs):
        call_n["i"] += 1
        if call_n["i"] == 1:
            return MagicMock(text="Sales by region.")
        raise RuntimeError("sampling boom")

    ctx = MagicMock()
    ctx.sample = flaky_sample

    job = registry.create(str(folder), db)
    asyncio.run(_run_index_job(job, None, True, None, ctx=ctx))

    fresh = VectorStore(db)
    # Atomic per-file commit: nothing written, entry still queued
    assert fresh.count_chunks() == 0
    assert fresh.pending_count() == 1


# --- Disabled state -------------------------------------------------------
#
# While spreadsheets are disabled, these tests must pass regardless. They
# cover the three behaviours that matter for the disable:
#   1. Default discovery skips spreadsheets.
#   2. Explicit `file_types` does NOT bypass the disable.
#   3. Pre-existing description chunks survive orphan cleanup on a re-run.


def test_csv_not_discovered_when_disabled(tmp_path):
    folder = tmp_path / "data"
    folder.mkdir()
    _write_csv(folder, "x.csv")
    db = str((tmp_path / "db").resolve())

    result = index_folder(str(folder), db_path=db)

    assert result["files_indexed"] == 0
    assert result["descriptions_queued"] == 0
    assert VectorStore(db).pending_count() == 0


def test_workbook_variants_not_discovered_when_disabled(tmp_path):
    folder = tmp_path / "data"
    folder.mkdir()
    _write_xlsx(folder, "a.xlsx", {"S": [["a"], [1]]})
    db = str((tmp_path / "db").resolve())

    result = index_folder(str(folder), db_path=db)

    assert result["files_indexed"] == 0
    assert result["descriptions_queued"] == 0


def test_explicit_file_types_does_not_bypass_disable(tmp_path):
    """Even when the caller passes spreadsheet extensions explicitly, the
    defensive filter in discover_files strips them."""
    folder = tmp_path / "data"
    folder.mkdir()
    _write_csv(folder, "x.csv")
    _write_xlsx(folder, "y.xlsx", {"S": [["a"], [1]]})
    db = str((tmp_path / "db").resolve())

    result = index_folder(
        str(folder),
        file_types=[".csv", ".xlsx", ".xlsm", ".xls"],
        db_path=db,
    )

    assert result["files_indexed"] == 0
    assert result["descriptions_queued"] == 0


def test_orphan_cleanup_preserves_existing_spreadsheet_chunks(tmp_path):
    """A description chunk written under a previous `enabled` run must NOT
    be wiped by orphan cleanup once spreadsheets are disabled — otherwise
    a single `index_folder` after the disable would silently delete every
    spreadsheet description chunk."""
    folder = tmp_path / "data"
    folder.mkdir()
    csv = _write_csv(folder, "old.csv")
    db = str((tmp_path / "db").resolve())
    rel = to_relative(str(csv), db)

    import numpy as np
    from server.store import EMBEDDING_DIM
    rng = np.random.RandomState(42)
    vec = rng.randn(EMBEDDING_DIM).astype(np.float32)
    vec = (vec / np.linalg.norm(vec)).tolist()
    store = VectorStore(db)
    store.ensure_schema()
    store.add_chunks([{
        "id": "desc_0",
        "text": "A CSV about sales orders.",
        "source_file": rel,
        "file_name": "old.csv",
        "file_type": ".csv",
        "folder_path": ".",
        "chunk_index": 0,
        "content_hash": "stable",
        "mtime_ns": 0,
        "file_size": 0,
        "vector": vec,
        "chunk_kind": "file_description",
        "sheet_name": None,
    }])
    assert store.count_chunks() == 1

    # Re-run index_folder. With .csv not in SUPPORTED_EXTENSIONS, orphan
    # cleanup must skip this row instead of deleting it.
    index_folder(str(folder), db_path=db)

    fresh = VectorStore(db)
    assert fresh.count_chunks() == 1, "existing description chunk was wiped"


@requires_spreadsheets_enabled
def test_auto_drainer_skipped_when_ctx_is_none(mock_embed_model, tmp_path):
    """Without a ctx, drain is skipped — entry stays queued for the LLM."""
    from server.main import _run_index_job
    from server.jobs import registry

    registry._jobs.clear(); registry._order.clear()
    registry._persist_path = None

    folder = tmp_path / "data"
    folder.mkdir()
    _write_csv(folder, "x.csv")
    db = str((tmp_path / "db").resolve())

    job = registry.create(str(folder), db)
    asyncio.run(_run_index_job(job, None, True, None, ctx=None))

    fresh = VectorStore(db)
    assert fresh.pending_count() == 1
    final = job.to_dict()
    assert final["result"]["descriptions_sampled"] == 0
