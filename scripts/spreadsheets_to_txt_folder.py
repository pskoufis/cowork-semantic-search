"""Batch-convert spreadsheets under a folder into plain-text ``.txt`` files.

Standalone tool — NOT wired into the indexer/server pipeline. It walks
``<input-folder>`` recursively, turns every spreadsheet it finds into tab-
separated text, and writes the result into ``<output-folder>`` mirroring the
input's subdirectory layout. The source tree is never modified.

Recognised extensions (case-insensitive): ``.csv .xlsx .xlsm .xls``.

Output layout — a workbook is a *container of sheets*, so it becomes a folder
of per-sheet text files; a CSV is a single sheet, so it becomes a single file::

    <in>/a/data.csv          ->  <out>/a/data.txt
    <in>/a/report.xlsx       ->  <out>/a/report/<sheet>.txt   (one per sheet)

Each ``.txt`` is capped in size. Rows are written until adding the next one
would cross ``--target-mb`` (default 10 MB), then a marker line is appended::

    ... [truncated: wrote N of M rows to stay under 10 MB]

Stopping at the 10 MB target keeps every output comfortably under the 20 MB
``--max-mb`` ceiling (so a 12 MB sheet is still trimmed). ``--max-mb`` is also
enforced as a hard cap: the header row plus the first data row are kept even
when they exceed the soft target, but never past the cap (a pathologically large
opening row yields fewer rows, possibly none). Rows are never split mid-row.

A structured JSONL run-log is written under ``<output>/_runlogs/`` (override
with ``--log-dir``); it records per-file status and drives an idempotent
re-run — a second invocation skips files already converted (output present,
source unchanged) and only (re)processes new or previously-failed ones. Pass
``--force`` to reconvert everything in place, ``--no-runlog`` to disable the
log, and ``--dry-run`` to preview what would be converted.

Usage:
    python scripts/spreadsheets_to_txt_folder.py <input-folder> <output-folder>
        [--target-mb MB] [--max-mb MB] [--dry-run]
        [--log-dir PATH] [--force] [--no-runlog]
"""

from __future__ import annotations

import argparse
import itertools
import os
import shutil
import sys
import tempfile
import time
from pathlib import Path
from typing import Iterable, Iterator

# Allow direct invocation (``python scripts/spreadsheets_to_txt_folder.py``) in
# addition to ``python -m scripts.spreadsheets_to_txt_folder`` by putting the
# repo root on sys.path before importing the sibling run-log helper.
_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from scripts._runlog import RunLog  # noqa: E402
from scripts import _runlog  # noqa: E402

# Every spreadsheet extension we read, mapped to the run-log ``kind`` tag.
SPREADSHEET_KINDS = {
    ".csv": "csv",
    ".xlsx": "xlsx",
    ".xlsm": "xlsm",
    ".xls": "xls",
}

DEFAULT_TARGET_MB = 10.0
DEFAULT_MAX_MB = 20.0

# Characters illegal in path components on common filesystems (Windows is the
# strict superset), replaced with ``_`` when deriving a file name from a sheet.
_BAD_NAME_CHARS = set('/\\:*?"<>|')


class UnreadableSpreadsheetError(Exception):
    """A spreadsheet could not be opened (corrupt, encrypted, non-tabular)."""


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    src = Path(args.input_folder).expanduser().resolve()
    dst = Path(args.output_folder).expanduser().resolve()

    if not src.exists():
        print(f"error: input folder not found: {src}", file=sys.stderr)
        return 2
    if not src.is_dir():
        print(f"error: input path is not a directory: {src}", file=sys.stderr)
        return 2
    if _is_inside(dst, src):
        print(
            f"error: output folder {dst} is inside input folder {src}; "
            "pick a target outside the input tree.",
            file=sys.stderr,
        )
        return 2
    if args.target_mb <= 0:
        print("error: --target-mb must be greater than 0", file=sys.stderr)
        return 2
    if args.max_mb <= args.target_mb:
        print(
            f"error: --max-mb ({args.max_mb:g}) must be greater than "
            f"--target-mb ({args.target_mb:g})",
            file=sys.stderr,
        )
        return 2

    target_bytes = int(args.target_mb * 1024 * 1024)
    max_bytes = int(args.max_mb * 1024 * 1024)

    files = _discover_spreadsheets(src)
    print(f"Found {len(files)} spreadsheet file(s) under {src}", file=sys.stderr)

    if args.dry_run:
        for f in files:
            rel = f.relative_to(src)
            print(f"would convert {rel} -> {_dry_run_target(f)}", file=sys.stderr)
        return 0

    dst.mkdir(parents=True, exist_ok=True)
    rl = RunLog(
        "spreadsheets_to_txt_folder",
        input_root=src,
        output_root=dst,
        argv=argv if argv is not None else sys.argv[1:],
        log_dir=args.log_dir,
        force=args.force,
        enabled=not args.no_runlog,
    )

    converted = skipped = failed = 0
    used_outputs: set[str] = set()
    started = time.monotonic()
    total = len(files)

    for i, path in enumerate(files, start=1):
        rel = path.relative_to(src)
        kind = SPREADSHEET_KINDS[path.suffix.lower()]

        # Idempotent re-run: a file already converted (output present, source
        # unchanged) is skipped. Reserve its output name so a different input
        # can't pick it.
        if (done := rl.done_output(path)) is not None:
            skipped += 1
            used_outputs.add(str(done))
            print(f"[{i}/{total}] skip (done) {rel}", file=sys.stderr)
            rl.record(path, kind=kind, output=done, status="skip")
            continue

        # Reuse the prior output path when reprocessing (force / changed) so we
        # overwrite in place rather than minting a duplicate ``name-1``.
        planned = rl.planned_output(path)
        if planned is not None:
            out_path = planned
            used_outputs.add(str(out_path))
        else:
            ext = ".txt" if kind == "csv" else ""  # "" => a directory
            out_path = _unique_out_path(dst / rel.parent, path.stem, ext, used_outputs)

        print(f"[{i}/{total}] converting {rel} -> {out_path.name}", file=sys.stderr)
        item_started = time.monotonic()
        try:
            _convert_one(
                path, out_path, kind=kind,
                target_bytes=target_bytes, max_bytes=max_bytes,
                target_mb=args.target_mb,
            )
        except Exception as exc:  # noqa: BLE001 — per-file isolation by design
            failed += 1
            print(f"error: failed to convert {path}: {exc}", file=sys.stderr)
            rl.record(path, kind=kind, output=out_path, status="fail", error=exc)
            continue
        converted += 1
        rl.record(path, kind=kind, output=out_path, status="ok",
                  duration_s=round(time.monotonic() - item_started, 3))

    elapsed = time.monotonic() - started
    skipped_part = f", skipped {skipped} (done)" if skipped else ""
    print(
        f"Converted {converted}, failed {failed}{skipped_part} "
        f"in {elapsed:.1f}s.",
        file=sys.stderr,
    )
    exit_code = 0 if failed == 0 else 1
    rl.finish(exit_code=exit_code)
    return exit_code


def _discover_spreadsheets(root: Path) -> list[Path]:
    """Every regular file under ``root`` whose extension is a spreadsheet
    format (case-insensitive), sorted deterministically by relative path."""
    results: list[Path] = []
    for path in root.rglob("*"):
        if path.is_file() and path.suffix.lower() in SPREADSHEET_KINDS:
            results.append(path)
    results.sort(key=lambda p: str(p.relative_to(root)))
    return results


def _unique_out_path(out_dir: Path, stem: str, ext: str, used: set[str]) -> Path:
    """Reserve a collision-free ``<out_dir>/<stem><ext>``.

    ``ext`` is ``".txt"`` for a CSV's single file or ``""`` for a workbook's
    per-sheet *directory*. Two inputs in the same output directory that share a
    stem (``book.xls`` + ``book.xlsx`` -> both want ``book/``) must not clobber
    each other, so the second and later get ``-1``, ``-2`` and a WARN. Both this
    run's reservations and entries already on disk (e.g. an untracked prior run)
    are avoided. Only used to *mint new* names — a tracked reprocess reuses
    ``planned_output`` and never lands here.
    """
    candidate = out_dir / f"{stem}{ext}"
    n = 0
    while str(candidate) in used or candidate.exists():
        n += 1
        candidate = out_dir / f"{stem}-{n}{ext}"
    if n:
        print(
            f"WARN: name collision for {stem!r} in {out_dir}; "
            f"writing {candidate.name}",
            file=sys.stderr,
        )
    used.add(str(candidate))
    return candidate


def _convert_one(
    src_path: Path, out_path: Path, *,
    kind: str, target_bytes: int, max_bytes: int, target_mb: float,
) -> None:
    """Convert one spreadsheet to text at ``out_path``.

    CSV -> a single ``.txt`` file. Workbook -> ``out_path`` is a directory
    holding one ``<sheet>.txt`` per non-empty sheet. On a reprocess the prior
    directory is cleared first so renamed/removed sheets don't leave orphans and
    re-runs never mint ``<sheet>-1.txt`` duplicates. Raises
    :class:`UnreadableSpreadsheetError` (and other read errors) on failure; the
    caller isolates it per-file.
    """
    if kind == "csv":
        _write_rows_to(
            out_path, _csv_rows(src_path),
            target_bytes=target_bytes, max_bytes=max_bytes,
            target_mb=target_mb, total_rows=None,
        )
        return

    # Workbook: rebuild the per-workbook directory from scratch. Prime the
    # sheet iterator first so an unreadable file raises (its workbook open is
    # lazy) *before* we touch out_path — a failed convert leaves no empty dir.
    # A readable-but-all-empty workbook still creates the dir, so it records as
    # done and won't be reprocessed every run.
    sheets = _iter_xls_sheets(src_path) if kind == "xls" else _iter_xlsx_sheets(src_path)
    try:
        first_sheet = next(sheets)
        sheets = itertools.chain([first_sheet], sheets)
    except StopIteration:
        sheets = iter(())  # no sheets at all (unusual); still create the dir

    if out_path.exists():
        shutil.rmtree(out_path)
    out_path.mkdir(parents=True, exist_ok=True)

    used_sheet_files: set[str] = set()
    written_sheets = 0
    for idx, name, rows, total_rows in sheets:
        row_iter = iter(rows)
        try:
            first = next(row_iter)
        except StopIteration:
            print(f"  note: skipping empty sheet {name!r}", file=sys.stderr)
            continue
        fname = _sanitize_sheet_name(name, idx)
        sheet_out = _unique_out_path(out_path, fname, ".txt", used_sheet_files)
        _write_rows_to(
            sheet_out, itertools.chain([first], row_iter),
            target_bytes=target_bytes, max_bytes=max_bytes,
            target_mb=target_mb, total_rows=total_rows,
        )
        written_sheets += 1

    if written_sheets == 0:
        print(f"  note: {src_path.name} had no non-empty sheets", file=sys.stderr)


def _write_rows_to(
    out_path: Path, rows: Iterable[Iterable[object]], *,
    target_bytes: int, max_bytes: int, target_mb: float, total_rows: int | None,
) -> None:
    """Write ``rows`` to ``out_path`` as tab-separated lines, truncating for size.

    Stops before any row that would push the file past ``target_bytes`` and
    appends a truncation marker. The first two rows (header + first data row) are
    kept even when they exceed the soft target, *but never past ``max_bytes``* —
    so the output is guaranteed under the hard cap, even for a pathologically
    large opening row (in which case fewer than two rows, possibly none, land).
    Written atomically: streamed to a temp file in the target directory and
    moved into place, so a failure mid-write leaves no partial output.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(
        dir=str(out_path.parent), prefix=".sstxt-", suffix=".tmp",
    )
    written = 0
    truncated = False
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="") as fh:
            size = 0
            for row in rows:
                line = "\t".join(_clean(_cell_to_str(c)) for c in row) + "\n"
                encoded = len(line.encode("utf-8"))
                # Soft target stops once we've kept header + first data row;
                # the hard cap is absolute and overrides that grace.
                over_target = written >= 2 and size + encoded > target_bytes
                over_cap = size + encoded > max_bytes
                if over_target or over_cap:
                    truncated = True
                    break
                fh.write(line)
                size += encoded
                written += 1
            if truncated:
                if total_rows is not None:
                    marker = (
                        f"... [truncated: wrote {written} of {total_rows} rows "
                        f"to stay under {target_mb:g} MB]\n"
                    )
                else:
                    marker = (
                        f"... [truncated: wrote {written} rows, more remain, "
                        f"to stay under {target_mb:g} MB]\n"
                    )
                fh.write(marker)
        shutil.move(tmp_name, str(out_path))
    except BaseException:
        try:
            os.unlink(tmp_name)
        except OSError:
            pass
        raise


def _cell_to_str(value: object) -> str:
    """Coerce a cell value to its text form.

    openpyxl returns Python objects (int, float, datetime, bool, None); xlrd
    surfaces every BIFF number as float even when the cell held an int. CSV
    cells are already strings. Stringify uniformly and normalise integral floats
    so "10" not "10.0" lands in the output. (Mirrors
    ``server/spreadsheets._cell_to_str``; copied so this script stays
    self-contained and never imports the server package.)
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _clean(cell: str) -> str:
    """Flatten tabs/newlines inside a cell to spaces so one row stays one line."""
    return cell.replace("\t", " ").replace("\r", " ").replace("\n", " ")


def _sanitize_sheet_name(name: str, idx: int) -> str:
    """Turn a sheet name into a safe file stem; fall back to ``sheet{idx}``."""
    cleaned = "".join(
        "_" if (ch in _BAD_NAME_CHARS or ord(ch) < 32) else ch for ch in name
    )
    cleaned = " ".join(cleaned.split())  # collapse runs of whitespace
    cleaned = cleaned.strip().strip(".")  # no trailing dot/space (Windows)
    return cleaned or f"sheet{idx}"


def _csv_rows(path: Path) -> Iterator[list[str]]:
    """Yield every row of a CSV as a list of strings (UTF-8, lossy on errors).

    The default ``csv`` field-size limit (128 KB) would raise on data exports
    with large free-text cells (email bodies, descriptions), failing the whole
    file; raise it to a generous, platform-safe bound first.
    """
    import csv

    _raise_csv_field_limit()
    with open(path, encoding="utf-8", errors="replace", newline="") as fh:
        for row in csv.reader(fh):
            yield row


def _raise_csv_field_limit() -> None:
    """Bump csv.field_size_limit as high as the platform's C long allows."""
    import csv

    limit = sys.maxsize
    while True:
        try:
            csv.field_size_limit(limit)
            return
        except OverflowError:
            limit //= 2


def _iter_xlsx_sheets(path: Path) -> Iterator[tuple[int, str, Iterator[tuple], int]]:
    """Yield ``(idx, sheet_name, rows_iter, total_rows)`` for an xlsx/xlsm.

    Uses openpyxl read-only + iter_rows so a sheet is never fully materialised.
    Each sheet's rows must be consumed before requesting the next.
    """
    import zipfile

    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError) as exc:
        raise UnreadableSpreadsheetError(str(exc)) from exc

    try:
        for idx, name in enumerate(wb.sheetnames):
            ws = wb[name]
            total = ws.max_row or 0

            def _rows(ws=ws):
                yield from ws.iter_rows(values_only=True)

            yield idx, name, _rows(), total
    finally:
        wb.close()


def _iter_xls_sheets(path: Path) -> Iterator[tuple[int, str, Iterator[list], int]]:
    """Yield ``(idx, sheet_name, rows_iter, total_rows)`` for a legacy .xls.

    Uses xlrd ``on_demand`` so sheets load lazily, and unloads each one after the
    caller has consumed its rows to keep the working set small.
    """
    import xlrd

    try:
        book = xlrd.open_workbook(str(path), on_demand=True)
    except Exception as exc:  # xlrd raises XLRDError/CompDocError/OSError/ValueError
        raise UnreadableSpreadsheetError(str(exc)) from exc

    try:
        for idx in range(book.nsheets):
            sheet = book.sheet_by_index(idx)
            total = sheet.nrows

            def _rows(sheet=sheet):
                for r in range(sheet.nrows):
                    yield [sheet.cell_value(r, c) for c in range(sheet.ncols)]

            yield idx, sheet.name, _rows(), total
            book.unload_sheet(idx)
    finally:
        book.release_resources()


def _dry_run_target(path: Path) -> str:
    """Human-readable description of what ``path`` would produce, for --dry-run."""
    kind = SPREADSHEET_KINDS[path.suffix.lower()]
    if kind == "csv":
        return f"{path.stem}.txt"
    try:
        if kind == "xls":
            count = sum(1 for _ in _iter_xls_sheets(path))
        else:
            count = sum(1 for _ in _iter_xlsx_sheets(path))
        return f"{path.stem}/ ({count} sheet(s))"
    except Exception as exc:  # noqa: BLE001 — preview only
        return f"{path.stem}/ (unreadable: {exc})"


def _is_inside(child: Path, parent: Path) -> bool:
    try:
        child.relative_to(parent)
    except ValueError:
        return False
    return True


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Recursively convert spreadsheets (.csv/.xlsx/.xlsm/.xls) under "
            "<input-folder> to tab-separated .txt in <output-folder>, one file "
            "per sheet, mirroring the input's subdirectory layout, capping each "
            "output's size by truncating rows. The input tree is never "
            "modified. Standalone — not part of the indexer."
        ),
    )
    parser.add_argument("input_folder", help="Folder to scan for spreadsheets")
    parser.add_argument("output_folder", help="Folder to write the .txt files into")
    parser.add_argument(
        "--target-mb",
        type=float,
        default=DEFAULT_TARGET_MB,
        help=f"Soft size target per output file in MB; stop adding rows before "
        f"crossing it (default: {DEFAULT_TARGET_MB:g}).",
    )
    parser.add_argument(
        "--max-mb",
        type=float,
        default=DEFAULT_MAX_MB,
        help=f"Hard size ceiling per output file in MB; must be greater than "
        f"--target-mb. Documents the guarantee that stopping at the target "
        f"keeps every file under this (default: {DEFAULT_MAX_MB:g}).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="List the files that would be converted and exit (writes nothing).",
    )
    _runlog.add_args(parser)
    return parser.parse_args(argv)


if __name__ == "__main__":
    raise SystemExit(main())
