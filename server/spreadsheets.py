"""Preview extraction and prompt building for CSV/XLSX files.

A "preview" is a structured dict the host LLM uses to describe a spreadsheet
without seeing every row. The shape is the same for CSV and XLSX so the
queue and prompt-builders can be uniform.
"""

import csv
import io
from pathlib import Path

SAMPLE_ROW_LIMIT = 10
MAX_DESCRIPTION_BYTES = 4096
SPREADSHEET_EXTENSIONS = {".csv", ".xlsx"}


def _infer_dtype(value: str) -> str:
    """Coarse dtype guess for a cell string. Used for prompt context only —
    not stored, not searchable."""
    v = value.strip()
    if not v:
        return "empty"
    try:
        float(v)
        return "number"
    except ValueError:
        return "string"


def _column_dtypes(sample_rows: list[list[str]], n_cols: int) -> list[str]:
    """Per-column dtype from the first non-empty sample row, fallback 'string'."""
    if not sample_rows or n_cols == 0:
        return []
    dtypes = ["string"] * n_cols
    for col in range(n_cols):
        for row in sample_rows:
            if col < len(row) and row[col].strip():
                dtypes[col] = _infer_dtype(row[col])
                break
    return dtypes


def _extract_csv_preview(file_path: Path) -> dict:
    """Build the preview for a single CSV file (a CSV is one 'sheet')."""
    raw = file_path.read_text(encoding="utf-8", errors="replace")
    if not raw.strip():
        return {
            "name": file_path.name,
            "headers": None,
            "first_row_as_data": True,
            "dtypes": [],
            "row_count": 0,
            "sample_rows": [],
        }

    # Sniffer needs a non-empty sample; cap to first 64KB so huge files don't
    # hold the whole text in a second buffer just for detection.
    sample = raw[:65536]
    try:
        has_header = csv.Sniffer().has_header(sample)
    except csv.Error:
        has_header = False

    reader = csv.reader(io.StringIO(raw))
    rows = list(reader)
    if not rows:
        return {
            "name": file_path.name,
            "headers": None,
            "first_row_as_data": True,
            "dtypes": [],
            "row_count": 0,
            "sample_rows": [],
        }

    if has_header:
        headers = rows[0]
        data_rows = rows[1:]
    else:
        headers = None
        data_rows = rows

    sample_rows = data_rows[:SAMPLE_ROW_LIMIT]
    n_cols = len(headers) if headers else (len(sample_rows[0]) if sample_rows else 0)
    dtypes = _column_dtypes(sample_rows, n_cols)

    return {
        "name": file_path.name,
        "headers": headers,
        "first_row_as_data": headers is None,
        "dtypes": dtypes,
        "row_count": len(data_rows),
        "sample_rows": sample_rows,
    }


class UnreadableSpreadsheetError(Exception):
    """Raised when a spreadsheet cannot be opened (corrupt, encrypted, etc).

    The caller treats this as a per-file failure: not enqueued, recorded in
    the file's error list. Encrypted/password-protected files are the
    common case but we deliberately don't try to distinguish — the LLM
    can't describe what we can't open, regardless of cause.
    """


def _cell_to_str(value) -> str:
    """Coerce a cell value to its preview-string form.

    openpyxl returns Python objects (int, float, datetime, bool, None);
    we stringify uniformly so the preview is JSON-encodable and consistent
    with the CSV path.
    """
    if value is None:
        return ""
    return str(value)


def _extract_xlsx_preview(file_path: Path) -> list[dict]:
    """Build per-sheet previews for a workbook.

    Uses openpyxl read_only mode + iter_rows so we never materialise the
    whole sheet. We pull one extra row beyond SAMPLE_ROW_LIMIT to compute
    row_count without iterating the rest (sheet.max_row from read-only mode
    is the workbook's recorded extent, which is reliable here).
    """
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
    import zipfile

    try:
        wb = openpyxl.load_workbook(
            filename=str(file_path), read_only=True, data_only=True,
        )
    except (InvalidFileException, zipfile.BadZipFile, KeyError) as exc:
        raise UnreadableSpreadsheetError(str(exc)) from exc

    previews: list[dict] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            collected: list[list[str]] = []
            for row in rows_iter:
                collected.append([_cell_to_str(c) for c in row])
                if len(collected) >= SAMPLE_ROW_LIMIT + 1:
                    break

            if not collected:
                previews.append({
                    "name": sheet_name,
                    "headers": None,
                    "first_row_as_data": True,
                    "dtypes": [],
                    "row_count": 0,
                    "sample_rows": [],
                })
                continue

            # Read-only mode: max_row is the workbook's stored extent. For
            # header-only sheets that's 1; for sheets with N data rows it's
            # N+1 (header + data).
            row_count_total = ws.max_row or len(collected)
            headers = collected[0]
            data_rows = collected[1:]
            sample = data_rows[:SAMPLE_ROW_LIMIT]
            data_count = max(row_count_total - 1, 0)
            n_cols = len(headers)
            dtypes = _column_dtypes(sample, n_cols)

            previews.append({
                "name": sheet_name,
                "headers": headers,
                "first_row_as_data": False,
                "dtypes": dtypes,
                "row_count": data_count,
                "sample_rows": sample,
            })
    finally:
        wb.close()

    return previews


def extract_preview(file_path: Path) -> dict:
    """Build the full preview dict for a spreadsheet.

    Returns {"type": "csv"|"xlsx", "sheets": [sheet_preview, ...]}.
    CSV has exactly one sheet (using the file name). XLSX has one entry
    per worksheet.

    Raises ValueError for unsupported extensions; UnreadableSpreadsheetError
    for files that can't be opened.
    """
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return {"type": "csv", "sheets": [_extract_csv_preview(file_path)]}
    if suffix == ".xlsx":
        return {"type": "xlsx", "sheets": _extract_xlsx_preview(file_path)}
    raise ValueError(f"Not a spreadsheet: {suffix}")


def needs_for_preview(preview: dict) -> list[str]:
    """Derive the queue's `needs` list from a preview.

    CSV → one file-level description (the CSV *is* its single sheet, so we
    skip the redundant sheet-level description). XLSX → one description
    per sheet plus a file-level rollup.
    """
    if preview["type"] == "csv":
        return ["file"]
    return [f"sheet:{s['name']}" for s in preview["sheets"]] + ["file"]


def _format_sheet_block(sheet: dict) -> str:
    """Render a single sheet's preview as a markdown-ish block for prompts."""
    headers = sheet["headers"]
    header_line = (
        f"Headers: {headers}" if headers
        else "Headers: (none detected — first row is data)"
    )
    if headers:
        n_cols = len(headers)
    elif sheet["sample_rows"]:
        n_cols = len(sheet["sample_rows"][0])
    else:
        n_cols = 0
    sample_lines = "\n".join(
        f"  {row}" for row in sheet["sample_rows"][:SAMPLE_ROW_LIMIT]
    ) or "  (no sample rows)"
    return (
        f"Sheet: {sheet['name']}\n"
        f"Rows: {sheet['row_count']}, Columns: {n_cols}\n"
        f"{header_line}\n"
        f"Dtypes: {sheet['dtypes']}\n"
        f"Sample:\n{sample_lines}"
    )


def build_sheet_prompt(preview: dict, sheet_name: str) -> str:
    """LLM prompt for a single sheet's description."""
    sheet = next(s for s in preview["sheets"] if s["name"] == sheet_name)
    return (
        "You are describing one sheet from a spreadsheet so that another "
        "system can later retrieve it by topic. Write a 1–3 sentence "
        "description of what this sheet contains and what it's used for. "
        "Focus on subject matter, not column-level mechanics. No preamble.\n\n"
        f"{_format_sheet_block(sheet)}"
    )


def build_file_prompt(preview: dict, sheet_descriptions: dict[str, str]) -> str:
    """LLM prompt for the file-level rollup description."""
    if preview["type"] == "csv":
        sheet = preview["sheets"][0]
        return (
            "You are describing a CSV file so that another system can later "
            "retrieve it by topic. Write a 1–3 sentence description of what "
            "this file contains and what it's used for. No preamble.\n\n"
            f"{_format_sheet_block(sheet)}"
        )
    desc_lines = "\n".join(
        f"- {name}: {desc}" for name, desc in sheet_descriptions.items()
    )
    return (
        "You are writing the file-level summary for a multi-sheet workbook "
        "so that another system can retrieve it by topic. Write a 2–4 "
        "sentence description of the workbook's purpose and the relationship "
        "between its sheets. No preamble.\n\n"
        f"Sheets and their descriptions:\n{desc_lines}"
    )
